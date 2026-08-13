"""Load reference lists still hosted in Google Sheets (gviz), phase-1 only."""

from __future__ import annotations

import json
import os
import re
from typing import Optional

import httpx

from matching import Client, enrich_client

MASTER_SHEET_ID = os.getenv("MASTER_SHEET_ID", "1Nk-LHFQSwXNDBBBq0tAj471ZE6zJu2C-soVFmjlBDO4")
EMPLOYEE_SHEET_ID = os.getenv("EMPLOYEE_SHEET_ID", "1dpZzyBAf9NwT2_XXsEN8443xfrEV5qeYxKjZhZLSw1c")
LAUNCHPAD_SHEET_ID = os.getenv("LAUNCHPAD_SHEET_ID", "1aXVaq39R6wsB0wgAAF_nxnuq6gqaRqmI1oHi5qqLpdI")
TIME_SHEET_ID = os.getenv("SHEET_ID", "1Aw7efJ_Z0VHiHH9iR2cm-HsP2OA_4hQCxeekSimSasY")
# Prefer local employee workbook (overrides outdated gviz directory when present).
_HERE = os.path.dirname(os.path.abspath(__file__))
EMPLOYEE_LIST_FILE = os.getenv(
    "EMPLOYEE_LIST_FILE",
    os.path.join(_HERE, "Digilatics_Employee_List.xlsx"),
)
CLIENTS_FILE = os.getenv(
    "CLIENTS_FILE",
    os.path.join(_HERE, "clients.json"),
)

_GVIZ_RE = re.compile(r"google\.visualization\.Query\.setResponse\(([\s\S]*?)\);")


def gviz_url(sheet_id: str, sheet: str) -> str:
    from urllib.parse import quote

    return (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq"
        f"?tqx=out:json&sheet={quote(sheet)}"
    )


def fetch_gviz_json(sheet_id: str, sheet: str, timeout: float = 60.0) -> dict:
    url = gviz_url(sheet_id, sheet)
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        res = client.get(url)
        res.raise_for_status()
        text = res.text
    m = _GVIZ_RE.search(text)
    if not m:
        raise ValueError(f"Could not parse gviz response for {sheet_id}/{sheet}")
    return json.loads(m.group(1))


def cell_val(row: dict, idx: int) -> Optional[str]:
    cells = row.get("c") or []
    if idx < 0 or idx >= len(cells) or not cells[idx]:
        return None
    v = cells[idx].get("v")
    if v is None:
        return None
    return str(v).strip()


def load_master_clients() -> tuple[list[Client], Optional[str]]:
    """Load client list — local clients.json first, Google Sheet fallback."""
    if CLIENTS_FILE and os.path.exists(CLIENTS_FILE):
        try:
            with open(CLIENTS_FILE, encoding="utf-8") as f:
                raw = json.load(f)
            clients: list[Client] = []
            for row in raw if isinstance(raw, list) else []:
                if not isinstance(row, dict):
                    continue
                name = str(row.get("name") or row.get("client") or "").strip()
                if not name or len(name) < 2:
                    continue
                aliases = row.get("aliases") or []
                domains = row.get("domains") or []
                if isinstance(aliases, str):
                    aliases = [s.strip() for s in aliases.split(",") if s.strip()]
                if isinstance(domains, str):
                    domains = [s.strip() for s in domains.split(",") if s.strip()]
                clients.append(
                    enrich_client(
                        name,
                        ",".join(str(a) for a in aliases),
                        ",".join(str(d) for d in domains),
                    )
                )
            if clients:
                return clients, None
        except Exception as e:
            local_err = str(e)
    else:
        local_err = f"missing {CLIENTS_FILE}"

    try:
        data = fetch_gviz_json(MASTER_SHEET_ID, "Sheet1")
        clients = []
        for r in data.get("table", {}).get("rows") or []:
            canonical = cell_val(r, 0)
            if not canonical or len(canonical) < 2:
                continue
            if canonical.lower() == "client name":
                continue
            aliases_raw = cell_val(r, 1) or ""
            domains_raw = cell_val(r, 2) or ""
            clients.append(enrich_client(canonical, aliases_raw, domains_raw))
        note = f"used_gviz_after_local:{local_err}" if local_err else None
        return clients, note
    except Exception as e:
        return [], f"local:{local_err}; gviz:{e}"


def load_employees_from_xlsx(path: str) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    from openpyxl import load_workbook

    email_name: dict[str, str] = {}
    email_fallback: dict[str, str] = {}
    email_team: dict[str, str] = {}
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c or "").strip().lower() for c in next(rows)]
    def idx(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return -1
    name_col = idx("employee name", "name")
    email_col = idx("email id", "email")
    fallback_col = idx("fallback client", "fallback")
    team_col = idx("team")
    if email_col < 0:
        wb.close()
        raise ValueError(f"email column missing in {path}")
    for row in rows:
        if not row or email_col >= len(row) or not row[email_col]:
            continue
        email = str(row[email_col]).strip().lower()
        name = str(row[name_col]).strip() if name_col >= 0 and name_col < len(row) and row[name_col] else None
        fallback = (
            str(row[fallback_col]).strip()
            if fallback_col >= 0 and fallback_col < len(row) and row[fallback_col]
            else "Digilatics"
        )
        team = str(row[team_col]).strip() if team_col >= 0 and team_col < len(row) and row[team_col] else None
        if name:
            email_name[email] = name
        email_fallback[email] = fallback
        if team:
            email_team[email] = team
    wb.close()
    return email_name, email_fallback, email_team


def _merge_users_json_employees(
    email_name: dict[str, str],
    email_fallback: dict[str, str],
    email_team: dict[str, str],
) -> None:
    """Overlay dashboard users.json (Settings → Add User) onto employee maps."""
    users_path = os.path.join(_HERE, "users.json")
    if not os.path.exists(users_path):
        return
    try:
        with open(users_path, encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return
    for email, cfg in (raw or {}).items():
        if not isinstance(cfg, dict):
            continue
        e = str(email).lower().strip()
        name = str(cfg.get("name") or "").strip()
        if name:
            # Prefer dashboard directory name over stale workbook short names
            email_name[e] = name
        email_fallback.setdefault(e, "Digilatics")
        teams = cfg.get("teams") or []
        if teams and str(teams[0]).strip():
            email_team.setdefault(e, str(teams[0]).strip())


def load_employees() -> tuple[dict[str, str], dict[str, str], dict[str, str], Optional[str]]:
    """Returns email_name_map, email_fallback_map, email_team_map, error."""
    # 1) Local workbook (source of truth when present)
    if EMPLOYEE_LIST_FILE and os.path.exists(EMPLOYEE_LIST_FILE):
        try:
            email_name, email_fallback, email_team = load_employees_from_xlsx(EMPLOYEE_LIST_FILE)
            _merge_users_json_employees(email_name, email_fallback, email_team)
            return email_name, email_fallback, email_team, None
        except Exception as e:
            xlsx_err = str(e)
    else:
        xlsx_err = f"missing {EMPLOYEE_LIST_FILE}"

    # 2) Fallback: gviz Employee Directory sheet
    email_name: dict[str, str] = {}
    email_fallback: dict[str, str] = {}
    email_team: dict[str, str] = {}
    try:
        data = fetch_gviz_json(EMPLOYEE_SHEET_ID, "Employee Directory")
        cols = [(c.get("label") or "").lower().strip() for c in data.get("table", {}).get("cols") or []]
        name_col = cols.index("employee name") if "employee name" in cols else -1
        email_col = cols.index("email id") if "email id" in cols else -1
        fallback_col = cols.index("fallback client") if "fallback client" in cols else -1
        team_col = cols.index("team") if "team" in cols else -1
        if email_col < 0:
            _merge_users_json_employees(email_name, email_fallback, email_team)
            return email_name, email_fallback, email_team, f"email id column missing (xlsx: {xlsx_err})"
        for r in data.get("table", {}).get("rows") or []:
            email = cell_val(r, email_col)
            if not email:
                continue
            email = email.lower()
            name = cell_val(r, name_col) if name_col >= 0 else None
            fallback = cell_val(r, fallback_col) if fallback_col >= 0 else None
            team = cell_val(r, team_col) if team_col >= 0 else None
            if name:
                email_name[email] = name
            email_fallback[email] = fallback or "Digilatics"
            if team:
                email_team[email] = team
        _merge_users_json_employees(email_name, email_fallback, email_team)
        return email_name, email_fallback, email_team, f"used_gviz_after_xlsx_error:{xlsx_err}"
    except Exception as e:
        _merge_users_json_employees(email_name, email_fallback, email_team)
        return email_name, email_fallback, email_team, f"xlsx:{xlsx_err}; gviz:{e}"


def digilatics_brand_names(email_fallback: dict[str, str]) -> list[str]:
    brands = sorted({(v or "").strip() for v in email_fallback.values() if v and str(v).strip()})
    if "Digilatics" not in brands:
        brands.append("Digilatics")
    return brands


def allowed_client_names(clients: list, email_fallback: Optional[dict[str, str]] = None) -> list[str]:
    names = [c.canonical for c in clients]
    if email_fallback:
        names.extend(digilatics_brand_names(email_fallback))
    # unique preserve order
    out, seen = [], set()
    for n in names:
        k = n.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(n)
    return out


def load_launchpad_subclients() -> list[str]:
    try:
        data = fetch_gviz_json(LAUNCHPAD_SHEET_ID, "subclients")
        out: list[str] = []
        for r in data.get("table", {}).get("rows") or []:
            name = cell_val(r, 0)
            if name and name.lower() != "launchpad subclients":
                out.append(name)
        return out
    except Exception:
        return []


def load_sheet_time_rows() -> list[dict]:
    """One-shot history import from the legacy time Sheet1 via gviz."""
    data = fetch_gviz_json(TIME_SHEET_ID, "Sheet1")
    cols = [c.get("label") or "" for c in data.get("table", {}).get("cols") or []]
    rows_out: list[dict] = []
    for r in data.get("table", {}).get("rows") or []:
        cells = r.get("c") or []
        mapped = {}
        for i, label in enumerate(cols):
            if i < len(cells) and cells[i] and cells[i].get("v") is not None:
                mapped[label] = cells[i].get("v")
            else:
                mapped[label] = ""
        entry_id = str(mapped.get("EntryId") or "").strip()
        if not entry_id:
            continue
        try:
            hours = float(mapped.get("Hours") or 0)
        except (TypeError, ValueError):
            hours = 0.0
        try:
            minutes = int(float(mapped.get("Minutes") or 0))
        except (TypeError, ValueError):
            minutes = 0
        rows_out.append(
            {
                "date": str(mapped.get("Date") or ""),
                "client": str(mapped.get("Client") or ""),
                "task": str(mapped.get("Task") or ""),
                "user": str(mapped.get("User") or ""),
                "hours": hours,
                "minutes": minutes,
                "source": str(mapped.get("Source") or ""),
                "url": str(mapped.get("URL") or ""),
                "entryId": entry_id,
                "space": str(mapped.get("Space") or ""),
                "team": str(mapped.get("Team") or ""),
                "matchVia": "sheet_import",
            }
        )
    return rows_out
