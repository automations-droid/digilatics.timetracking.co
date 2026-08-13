#!/usr/bin/env python3
"""Import Digilatics Time Reports .xlsx (Sheet1) into Postgres time_entries."""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")


def looks_entry_id(v) -> bool:
    s = str(v or "").strip()
    if not s or s.lower() in {"unknown", "none", "null", "-"}:
        return False
    if s.startswith("http"):
        return False
    if s.isdigit() and len(s) >= 8:
        return True
    if "_" in s and len(s) >= 10:
        return True
    if "@" in s and len(s) >= 10:
        return True
    return False


def looks_url(v) -> bool:
    return str(v or "").strip().lower().startswith("http")


def looks_source(v) -> bool:
    return str(v or "").strip().lower() in {
        "clickup",
        "meeting",
        "google_task",
        "task",
        "meet",
    }


def norm_source(v: str) -> str:
    s = (v or "").strip().lower()
    if s in {"meet", "meeting"}:
        return "meeting"
    if s in {"task", "clickup"}:
        return "clickup"
    if s == "google_task":
        return "google_task"
    return s


def infer_source(entry_id: str, url: str = "") -> str:
    u = (url or "").lower()
    e = entry_id or ""
    if "clickup.com" in u:
        return "clickup"
    if "@" in e or ("_" in e and not e.isdigit()):
        return "meeting"
    if e.isdigit():
        return "clickup"
    return "clickup"


def parse_rows(path: Path) -> tuple[list[dict], dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        raise RuntimeError(f"Sheet1 not found in {path}")
    ws = wb["Sheet1"]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        raise RuntimeError("Empty workbook")

    modes: Counter[str] = Counter()
    skipped: Counter[str] = Counter()
    out: list[dict] = []

    for row in it:
        cells = list(row) + [None] * 12
        d, client, task, user, hours = cells[0], cells[1], cells[2], cells[3], cells[4]
        c5, c6, c7, c8, c9 = cells[5], cells[6], cells[7], cells[8], cells[9]

        if isinstance(d, datetime):
            date_s = d.strftime("%Y-%m-%d")
        elif isinstance(d, date):
            date_s = d.isoformat()
        else:
            date_s = str(d or "").strip()[:10]
            if len(date_s) < 8:
                skipped["bad_date"] += 1
                continue

        try:
            h = float(hours)
        except (TypeError, ValueError):
            skipped["bad_hours"] += 1
            continue
        if h <= 0:
            skipped["nonpos_hours"] += 1
            continue
        if not user or not str(user).strip():
            skipped["no_user"] += 1
            continue

        source = ""
        url = ""
        entry_id = ""
        space = ""
        team = ""

        if looks_source(c5):
            source = norm_source(str(c5))
            url = str(c6 or "")
            entry_id = str(c7 or "").strip()
            space = str(c8 or "")
            team = str(c9 or "")
            modes["A_normal"] += 1
        elif looks_entry_id(c6) and not looks_source(c5) and not looks_url(c5):
            entry_id = str(c6).strip()
            t7 = str(c7 or "").strip()
            t8 = str(c8 or "").strip()
            t9 = str(c9 or "").strip()
            if t7.lower() in {"unknown", "-", ""} and t8:
                space, team = t8, t9 or "Unknown"
            elif t7:
                team, space = t7, t8 or "-"
            else:
                space, team = t8 or "-", t9 or "Unknown"
            source = infer_source(entry_id)
            modes["B_shifted_no_source"] += 1
        elif looks_url(c5):
            url = str(c5).strip()
            if looks_entry_id(c6):
                entry_id = str(c6).strip()
                space, team = str(c7 or ""), str(c8 or "")
            elif looks_entry_id(c7):
                entry_id = str(c7).strip()
                space, team = str(c8 or ""), str(c9 or "")
            source = infer_source(entry_id, url)
            modes["C_url_in_source"] += 1
        elif looks_entry_id(c5):
            entry_id = str(c5).strip()
            space, team = str(c6 or ""), str(c7 or "")
            source = infer_source(entry_id)
            modes["D_entry_in_source"] += 1
        else:
            found = None
            for c in cells[5:12]:
                if looks_entry_id(c):
                    found = c
                    break
                if looks_url(c) and not url:
                    url = str(c)
            if found is not None:
                entry_id = str(found).strip()
                source = infer_source(entry_id, url)
                modes["E_scan"] += 1
            else:
                # Keep rows with hours/user/date even without upstream id
                entry_id = f"xlsx:{date_s}:{str(user).strip()}:{str(task or '').strip()}:{h}"
                source = "clickup"
                space = str(c8 or c7 or "-")
                team = str(c9 or c7 or "Unknown")
                modes["F_synthetic"] += 1

        if not entry_id:
            skipped["empty_entryid"] += 1
            continue

        out.append(
            {
                "date": date_s,
                "client": str(client or ""),
                "task": str(task or ""),
                "user": str(user).strip(),
                "hours": h,
                "minutes": int(round(h * 60)),
                "source": source or "clickup",
                "url": url or "",
                "entryId": entry_id,
                "space": space or "-",
                "team": team or "Unknown",
                "matchVia": "xlsx_import",
            }
        )

    wb.close()
    stats = {
        "modes": dict(modes),
        "skipped": dict(skipped),
        "rows": len(out),
        "uniqueEntryIds": len({r["entryId"] for r in out}),
        "sources": dict(Counter(r["source"] for r in out)),
        "dateMin": min((r["date"] for r in out), default=None),
        "dateMax": max((r["date"] for r in out), default=None),
    }
    return out, stats


def main() -> int:
    p = argparse.ArgumentParser(description="Import time report xlsx into Postgres")
    p.add_argument(
        "xlsx",
        nargs="?",
        default=str(Path.home() / "Downloads" / "Digilatics Time Reports (4).xlsx"),
        help="Path to Digilatics Time Reports xlsx",
    )
    args = p.parse_args()
    path = Path(args.xlsx).expanduser().resolve()
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    print(f"Parsing {path} …", flush=True)
    rows, stats = parse_rows(path)
    print(stats, flush=True)
    if not rows:
        print("No rows to import", file=sys.stderr)
        return 1

    # Postgres ON CONFLICT cannot touch the same entry_id twice in one statement
    deduped: dict[str, dict] = {}
    for r in rows:
        deduped[str(r["entryId"])] = r
    rows = list(deduped.values())
    stats["rowsAfterDedupe"] = len(rows)
    print(f"After dedupe: {len(rows)} rows", flush=True)

    from db import count_entries, get_session, init_db, record_sync_run, upsert_entries

    init_db()
    session = get_session()
    started = datetime.now(timezone.utc)
    try:
        before = count_entries(session)
        print(f"DB before: {before}. Upserting {len(rows)} rows…", flush=True)
        inserted = upsert_entries(session, rows)
        after = count_entries(session)
        debug = {
            "file": str(path),
            "parse": stats,
            "rowsUpserted": inserted,
            "dbCountBefore": before,
            "dbCountAfter": after,
        }
        record_sync_run(
            session,
            job="xlsx_import",
            rows_inserted=inserted,
            status="ok",
            debug=debug,
            started_at=started,
        )
        print(debug, flush=True)
        print("OK", flush=True)
        return 0
    except Exception as e:
        session.rollback()
        record_sync_run(
            session,
            job="xlsx_import",
            rows_inserted=0,
            status="error",
            debug={"error": str(e), "file": str(path)},
            started_at=started,
        )
        raise
    finally:
        session.close()


if __name__ == "__main__":
    raise SystemExit(main())
